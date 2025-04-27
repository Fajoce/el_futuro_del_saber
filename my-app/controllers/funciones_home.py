
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string
import pdb
from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio


import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file


def procesar_form_estudiante(tipo_doc, num_doc, nombres, apellidos,fecha_nacimiento, grado, direccion, email, tel_fijo):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                sql = """
                INSERT INTO estudiante 
                (tipo_doc, num_doc, nombres, apellidos,fecha_nacimiento, grado, direccion, email, tel_fijo) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                valores = (tipo_doc, num_doc, nombres, apellidos,fecha_nacimiento, grado, direccion, email, tel_fijo)
                mycursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                resultado_insert = mycursor.rowcount
                return resultado_insert
    except Exception as e:
        print(f"Error en el Insert students: {e}")
        return []

# Lista de Estudiantes
def sql_lista_estudiantesBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id_estudiante,tipo_doc, num_doc,nombres,apellidos,fecha_nacimiento,grado,ciudad_residencia, direccion, email,tel_fijo,celular,nombre_acudiente FROM estudiante"
                  
                cursor.execute(querySQL,)
                estudiantesBD = cursor.fetchall()
        return estudiantesBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_estudiantesBD: {e}")
        return None


# Detalles del Estudiante
def sql_detalles_estudiantesBD(id_estudiante):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_estudiante,
                        e.tipo_doc, 
                        e.num_doc,
                        e.nombres,
                        e.apellidos,
                        e.fecha_nacimiento,
                        e.grado,
                        e.ciudad_residencia,
                        e.direccion,
                        e.email,
                        e.tel_fijo,
                        e.celular,
                        e.nombre_acudiente
                    FROM estudiante AS e
                    WHERE id_estudiante =%s
                    ORDER BY e.id_estudiante DESC
                    """)
                cursor.execute(querySQL, (id_estudiante,))
                estudiantesBD = cursor.fetchone()
        return estudiantesBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_estudiantesBD: {e}")
        return None


# Funcion Estudiantes Informe (Reporte)
def estudiantesReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT e.id_estudiante, e.tipo_doc, e.num_doc, e.nombres,e.apellidos, e.tel_fijo,e.grado,e.email,fecha_nacimiento,direccion  FROM estudiante AS e  ORDER BY e.id_estudiante DESC"
                cursor.execute(querySQL,)
                estudiantesBD = cursor.fetchall()
        return estudiantesBD
    except Exception as e:
        print(
            f"Error en la función estudiantesReporte: {e}")
        return None


def generarReporteExcel():
    dataEstudiantes = estudiantesReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("id_estudiante","tipo_doc","num_doc","nombres", "apellidos", "tel_fijo", "email", "grado","fecha_nacimiento","direccion")

    hoja.append(cabeceraExcel)

    # Agregar los registros a la hoja
    for registro in dataEstudiantes:
        id_estudiante = registro['id_estudiante']
        tipo_doc = registro['tipo_doc']
        num_doc = registro['num_doc']
        nombres = registro['nombres']
        apellidos = registro['apellidos']
        tel_fijo = registro['tel_fijo']
        email = registro['email']
        grado = registro['grado']
        fecha_nacimiento = registro['fecha_nacimiento']
        direccion = registro['direccion']

        # Agregar los valores a la hoja
        hoja.append((id_estudiante,tipo_doc, num_doc,nombres, apellidos, tel_fijo, email, grado,fecha_nacimiento,direccion))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 10 # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
         

    fecha_actual = datetime.datetime.now()
    archivoExcel = "Reporte_estudiantes_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


def buscarEstudianteBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                        e.id_estudiante,
                        e.tipo_doc, 
                        e.num_doc,
                        e.nombres,
                        e.apellidos,
                        e.fecha_nacimiento,
                        e.grado
                        e.ciudad_residencia,
                        e.direccion,
                        e.email,
                        e.tel_fijo,
                        e.celular,
                        e.nombre_acudiente
                        FROM estudiante AS e
                        WHERE e.nombres LIKE %s 
                        ORDER BY e.id_estudiante DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarEstudianteBD: {e}")
        return []


def buscarEstudianteById(id_estudiante):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = "SELECT id_estudiante,tipo_doc, num_doc,nombres,apellidos,fecha_nacimiento,grado, ciudad_residencia,direccion, email,tel_fijo,celular,nombre_acudiente  FROM estudiante  WHERE id_estudiante =%s LIMIT 1"
                    
                mycursor.execute(querySQL, (id_estudiante,))
                estudiante = mycursor.fetchone()
                return estudiante

    except Exception as e:
        print(f"Ocurrió un error en def buscarEstudianteById: {e}")
        return []

 
def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                tipo_doc = data.form['tipo_doc']
                num_doc = data.form['num_doc']
                nombres = data.form['nombres']
                apellidos = data.form['apellidos']
                direccion = data.form['direccion']
                tel_fijo = data.form['tel_fijo']
                email = data.form['email']
                grado = data.form['grado']
                id_estudiante = data.form['id_estudiante']

                query_sql = """
                    UPDATE estudiante
                    SET 
                        tipo_doc = %s,
                        num_doc = %s,
                        nombres = %s,
                        apellidos = %s,
                        direccion = %s,
                        tel_fijo = %s,
                        email = %s,
                        grado = %s
                    WHERE id_estudiante = %s
                """

                params = [
                    tipo_doc, num_doc, nombres, apellidos,
                    direccion, tel_fijo, email, grado,
                    id_estudiante
                ]

                cursor.execute(query_sql, params)
                conexion_MySQLdb.commit()
                print(f"Filas afectadas: {cursor.rowcount}")

        return cursor.rowcount
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None
    
# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []


# Eliminar uEstudiante
def RemoverEstudiante(id_estudiante):
    try:
        
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM estudiante WHERE id_estudiante=%s"
                cursor.execute(querySQL, (id_estudiante,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
                
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEstudiante : {e}")
        return []


# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []

def obtener_estudiantes_paginados(offset=0, limit=10):
    try:
        with connectionBD() as conexion:
            with conexion.cursor(dictionary=True) as cursor:
                query = "SELECT * FROM estudiante LIMIT %s OFFSET %s"
                cursor.execute(query, (limit, offset))
                return cursor.fetchall()
    except Exception as e:
        print(f"Error: {e}")
        return []

def contar_estudiantes():
    try:
        with connectionBD() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM estudiante")
                total = cursor.fetchone()[0]
                return total
    except Exception as e:
        print(f"Error: {e}")
        return 0
